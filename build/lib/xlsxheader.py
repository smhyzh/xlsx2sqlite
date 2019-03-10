#!/usr/bin/env python3
#-*- coding:utf-8 -*-
# Author:smh

class XlsxHeader():
    '''Find the table header of a xlsx worksheet.
    1.table header include table title and table items.
    2.table title shoule be a merge cell and length equal with max_column.
    3.table items not include merge cell.
    4.If find a not merge row has the same length with max_column,use it as table items.
    '''
    def __init__(self,work_sheet, forcase_line=10):
        '''
        '''
        self._ws = work_sheet
        if self._ws.max_row < forcase_line:
            self._forcase_row = self._ws.max_row
        else:
            self._forcase_row = forcase_line

    def inMergeCell(self,merge_cell,cell_row,cell_col):
        '''Judge a cell in the merge_cell.
        merge_cell: the any value of merged_cells.
        cell_row: the row index of the cell to test.
        cell_col: the col index of the cell to test.
        '''
        if merge_cell.min_row <= cell_row <= merge_cell.max_row and merge_cell.min_col <= cell_col <= merge_cell.max_col:
            return True
        else:
            return False

    def hasLineMergeCell(self, index):
        '''
        '''
        merge_list = self._ws.merged_cells

        for merge in merge_list:
            if merge.min_row <= index <= merge.max_row:
                return True
        return False

    def getCellValue(self,row,col):
        '''Get the cell value by the row and col.
        If the row,col in a merge_cell,will return the left-top cell's value.
        '''
        #For merge cells,in windows the methond merge_cell.value will
        #raise a exception.  AttributeError: 'MergedCell' object has 
        # no attribute 'value'

        #get merge list.
        merge_list = self._ws.merged_cells

        for merge in merge_list:
            if self.inMergeCell(merge,row,col):
                return self._ws.cell(merge.min_row,merge.min_col).value
            else:
                return self._ws.cell(row,col).value
        
    def getLine(self,index):
        '''Return the value of the row specified by Index.
        index:the row position. Start by 1.
        '''
        if index < 0 :
            raise IndexError('index out of range.')

        ret_list = []

        for col in range(1,len(self._ws[index])+1):
            ret_list.append(self.getCellValue(index,col))

        return ret_list

    def getLineRaw(self,index):
        '''Return the row specified by Index.
        index:the row position. Start by 1.
        '''
        if index < 0 :
            raise IndexError('index out of range.')
        return self._ws[index]

    def countValidItem(self,line_data):
        '''Count the number of valide item in a row.
        line_data:the data of line,get by getLine().
        '''
        count = 0
        for item in line_data:
            if item:
                count+=1
        return count

    def getTableItems(self):
        '''Find the first row which have same length with max_column not merged to be table items.
        '''
        for row in range(1,self._forcase_row+1):
            if self.hasLineMergeCell(row):
                continue
            if self.countValidItem(self.getLine(row)) == self._ws.max_column:
                return row
        raise RuntimeError('Not found the table items.')

if __name__ == "__main__":
    from openpyxl import load_workbook
    wb = load_workbook('test.xlsx')
    label = 'Sheet1'
    ws = wb[label]
    xhdr = XlsxHeader(ws)
    hdr = xhdr.getTableItems()
    print(hdr)