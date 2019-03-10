#!/usr/bin/env python3
#-*- coding:utf-8 -*-
# Author:smh

from dbopt import MyDateBase
from openpyxl import load_workbook
from xlsxheader import XlsxHeader


class XlsxToSqlite():
    '''
    '''
    def __init__(self,work_book):
        '''
        '''
        self._wb = work_book

    def convertToSqlite(self,table_label,table_name=None,sqlite_db_path=None):
        '''
        Convert xlsx to sqlite database will lost cells style information.
        '''
        if table_label not in self._wb.sheetnames:
            raise ValueError('table_label:{} not found.'.format(table_label))

        if not sqlite_db_path:
            sqlite_db_path = 'test.db'

        if not table_name:
            table_name = table_label

        ws = self._wb[table_label]

        xhdr = XlsxHeader(ws)
        try:
            item_index = xhdr.getTableItems()
        except RuntimeError:
            print('not found table items.\nconvert to sqlite failed!')
        
        items_list = [ item for item in xhdr.getLine(item_index)]

        mdbs = MyDateBase(sqlite_db_path)
        with mdbs as mdb:
            sql_str = mdb.constructHeaderSQL(table_name,items_list)
            print(sql_str)
            mdb.runSQL(sql_str)
            for every_row in list(ws.rows)[item_index+1:]:
                row_data = [item.value for item in every_row]
                sql_str = mdb.insertRowSQL(row_data)
                mdb.runSQL(sql_str,row_data)


if __name__ == "__main__":
    test_path = 'test.xlsx'
    wb = load_workbook(test_path)
    test_label = 'Sheet1'
#     ws = wb[test_label]
#     xhdr = XlsxHeader(ws)
#     for index in range(1,10):
#         item_list = xhdr.getLine(index)
#         for item in item_list:
#             print(item.value,end='\t')
#         print('')
#     index = xhdr.getTableItems()
#     print('items:{}'.format(index))
#     item_list = xhdr.getLine(index)
#     for item in item_list:
#         print(item.value,end='\t')
#     print('') 
    myxl = XlsxToSqlite(wb)
    myxl.convertToSqlite(test_label)
